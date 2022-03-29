# Run as follows:
# etm.py API_KEY
# where API_KEY is the API key for ETM product

import time
import requests
from openpyxl import load_workbook, Workbook
import os
import sys

api_key = sys.argv[1]
wb = load_workbook('ETMinput.xlsx')
ws = wb['API']
rows = [tuple(row[0:13]) for row in ws.iter_rows(min_row=2, max_row=2)]

for row in rows:
    # get list of existing tasks
    status = requests.get(url='https://demo.elseviertextmining.com/api/export/task?apikey=%s' % (api_key))
    old_status = status.json()
    old_jobs = [x['taskId'] for x in old_status]
    
    # create new tasks and get their IDs
    n = row[0].value
    disease = row[1].value
    urls = [x.value for x in row[2:]]
    for job_request_url in urls:
        job_request = requests.get(url=job_request_url)
    status = requests.get(url='https://demo.elseviertextmining.com/api/export/task?apikey=%s' % (api_key))
    new_status = status.json()
    new_jobs = [x['taskId'] for x in new_status if x['taskId'] not in old_jobs]
    
    # wait for new tasks to finish
    now = time.time()
    canceled_jobs = []
    completed = False
    while not completed:
        for job_id in canceled_jobs:
            if job_id in new_jobs:
                new_jobs.remove(job_id)
        time.sleep(5)
        completed = True
        for job_id in new_jobs:
            status = requests.get(url='https://demo.elseviertextmining.com/api/export/task/%s?apikey=%s' % (job_id, api_key))
            current_status = status.json()
            if current_status['state'] not in ['COMPLETED', 'CANCELED', 'FAILED']:
                completed = False
                if time.time() - now > 300:
                    requests.delete(url='https://demo.elseviertextmining.com/api/export/task/%s?apikey=%s' % (job_id, api_key))
            if current_status['state'] in ['CANCELED', 'FAILED']:
                canceled_jobs.append(job_id)
    
    # download files
    files = []
    for job_id in new_jobs:
        if job_id in canceled_jobs:
            continue
        file_request_url = 'https://demo.elseviertextmining.com/api/export/task/%s/result?apikey=%s' % (job_id, api_key)
        file_request = requests.get(url=file_request_url)
        filename = '%s.xlsx' % (job_id)
        print(filename)
        with open(filename, mode='wb') as f:
            f.write(file_request.content)
        files.append(filename)
    
    out_book = Workbook()
    out_sheet = out_book.active
    header_rows = 9
    top_count = 25
    current_row = 0
    for filename in files:
        content = load_workbook(filename)
        sheet = content['Query Summary']
        row_count = len(list(sheet.rows))
        col_count = len(list(sheet.columns))
        acquire_rows = min(header_rows + top_count, row_count)
        for row_number in range(1, acquire_rows+1):
            for col_number in range(1, col_count+1):
                out_sheet.cell(row_number + current_row, col_number).value = sheet.cell(row_number, col_number).value
        out_sheet.cell(current_row + 1, 2).value = filename
        current_row += acquire_rows + 1
        os.replace(filename, './cache/%s' % (filename))
    out_book.save('%s_{:0>3d}.xlsx'.format(n) % (disease))
    print('done %s' % (str(n) + '_' + disease))
